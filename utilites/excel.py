import openpyxl
from openpyxl.drawing.image import Image
from datetime import datetime


def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_row)


def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_column)


def readData(file,sheetName,rownum,columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rownum, column=columnnum).value


def writeData(file,sheetName,rownum,columnnum,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=rownum, column=columnnum).value = data
    workbook.save(file)


def writeImg(file, sheetName, rownum, columnum, img_path):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    img = Image(img_path)
    img.height = 800
    img.width = 1200
    sheet.cell(row=rownum, column=columnum).value = img
    workbook.save(file)


def excel_result(file, sheet_name, row, result):
    workbook = openpyxl.load_workbook(file)
    run_date = datetime.now().strftime('%Y-%m-%d')
    run_time = datetime.now().strftime('%H:%M:%S')
    # Use get_sheet_by_name to select the desired sheet
    sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(
        sheet_name)

    # Find the last column index in the header row
    last_column = sheet.max_column

    # Update values for each row
    sheet.cell(row=row, column=last_column - 2, value=run_date)
    sheet.cell(row=row, column=last_column - 1, value=run_time)
    sheet.cell(row=row, column=last_column, value=result)
    workbook.save(file)



